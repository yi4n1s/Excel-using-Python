from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = load_workbook('regions.xlsx')
ws = wb.active

cell_range = ws['A1':'C1']
#print(cell_range)
col_c = ws['C']
#print(col_c)
col_range = ws['A':'C']
#print(col_range)
row_range = ws[1:5]
#print(row_range)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
	for cell in row:
		print(cell)