import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

df_1 = pd.read_excel('shifts.xlsx', sheet_name='Sheet')
df_2 = pd.read_excel('shifts.xlsx', sheet_name='Sheet1')
df_3 = pd.read_excel('shift_3.xlsx', sheet_name='Sheet1')
df_all = pd.concat([df_1, df_2, df_3], sort=False)
#print(df_all)
#print(df_all.loc[50])
#print(df_all.groupby(['Shift']).mean()['Units Sold'])

#to_excel = df_all.to_excel('all_shifts.xlsx', index=None)

wb = load_workbook('all_shifts.xlsx')
ws = wb.active

total_col = ws['G1']
total_col.font = Font(bold=True)
total_col.value = 'Total'

e_col, f_col = ['E', 'F']
for row in range(2,300):
	result_cell = 'G{}'.format(row)
	e_value = ws[e_col + str(row)].value
	f_value = ws[f_col + str(row)].value
	ws[result_cell] = e_value * f_value

wb.save('totaled.xlsx')